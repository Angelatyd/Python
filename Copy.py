import openpyxl

wb1 = openpyxl.load_workbook(r'C:\Users\atian1\PycharmProjects\PythonTools\dashboard_sum.xlsx',keep_vba=True)
wb2 = openpyxl.load_workbook(r'C:\Users\atian1\PycharmProjects\PythonTools\test.xlsm',keep_vba=True)

sh1 = wb1['pyt']
sh2 = wb2['Sheet1']
j = sh1.max_row
for i in range(1, 6):
    for j in range(2,j+1):
        sh2.cell(row=j, column=i).value = sh1.cell(row=j, column=i+1).value

wb2.save(r'C:\Users\atian1\PycharmProjects\PythonTools\test.xlsm')